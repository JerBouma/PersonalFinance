from openpyxl import load_workbook
import pandas as pd
import os


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def data_processor(file):
    data = pd.read_csv(file)

    debit_credit = data.columns[5]
    amount = data.columns[6]

    data[amount] = data[amount].apply(
        lambda x: x.replace(',', '.'))
    data[amount] = pd.to_numeric(data[amount])
    values = []

    for row, value in data.iterrows():
        if value[debit_credit] in ('Debit', 'Af'):
            values.append(-value[amount])
        elif value[debit_credit] in ('Credit', 'Bij'):
            values.append(value[amount])

    data[amount] = values

    data = data.set_index(data[data.columns[0]])
    data = data.drop(
        [data.columns[0], data.columns[2],
         data.columns[3], data.columns[4],
         data.columns[5], data.columns[7]], axis=1)

    return data


def load_personal_input(file):
    wb = load_workbook(file)
    data = {}

    for category in wb.sheetnames:
        data[category] = []
        for cell in wb[category]['A']:
            cell_value = cell.value.lower()
            data[category].append(cell_value)

    return data


def category_selector(input, bank_data):
    categories = []
    notifications = bank_data.columns[2]
    description = bank_data.columns[0]

    for row, value in bank_data.iterrows():
        category_decision = "Other"
        for category in input.keys():
            if category_decision == "Other":
                for item in input[category]:
                    if item in value[notifications].lower():
                        category_decision = category
                        continue
                    elif item in value[description].lower():
                        category_decision = category
                        continue
        categories.append(category_decision)

    bank_data["Categories"] = categories

    return bank_data
