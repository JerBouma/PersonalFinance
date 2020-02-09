from openpyxl import Workbook, load_workbook
import pandas as pd
import os


def excel_categories(input, output, bank_data):
    categories = list(input.keys())
    categories.append("Other")

    sheet_name = 'Categories.xlsx'
    folder = output

    book = Workbook()
    writer = pd.ExcelWriter(folder + "\\" + sheet_name,
                            engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for category in categories:
        data = bank_data[bank_data["Categories"] == category].drop("Categories",
                                                                   axis=1)

        data.to_excel(writer, category)

        sheet = book[category]
        sheet.column_dimensions['B'].width = len(bank_data.columns[0]) * 2.0
        sheet.column_dimensions['C'].width = len(bank_data.columns[1]) * 1.5
        sheet.column_dimensions['D'].width = len(bank_data.columns[1]) * 10
        sheet.sheet_view.showGridLines = False

    try:
        book.remove_sheet(book['Sheet'])
    except KeyError:
        None

    writer.save()


def excel_bank_data_years(output, bank_data):
    bank_data['Day'] = bank_data.index.astype(str).str[6:8]
    bank_data['Month'] = bank_data.index.astype(str).str[4:6]
    bank_data['Year'] = bank_data.index.astype(str).str[:4]

    folder = output + '//Banksheets'

    try:
        os.mkdir(folder)

    except FileExistsError:
        None

    for year in bank_data['Year'].unique():
        book = Workbook()
        sheet_name = "Banksheet " + year + ".xlsx"

        writer = pd.ExcelWriter(folder + "\\" + sheet_name,
                                engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        months = bank_data['Month'][bank_data['Year'] == year]

        for month in months.unique():
            data = bank_data[(bank_data['Year'] == year)
                             & (bank_data['Month'] == month)]
            data = data.sort_values(by='Day')
            data = data.set_index(data['Day'])
            data = data.drop(['Day', 'Month', 'Year'], axis=1)

            data.to_excel(writer, month)

            sheet = book[month]
            sheet.column_dimensions['B'].width = len(data.columns[0]) * 2.0
            sheet.column_dimensions['C'].width = len(data.columns[1]) * 1.5
            sheet.column_dimensions['D'].width = len(data.columns[2]) * 10
            sheet.column_dimensions['E'].width = len(data.columns[3]) * 2.0
            sheet.sheet_view.showGridLines = False

        try:
            book.remove_sheet(book['Sheet'])
        except KeyError:
            None

        writer.save()


def excel_report(input, output, bank_data):
    bank_data_final = {}
    total_list = {}

    bank_data['Month'] = bank_data.index.astype(str).str[4:6]
    bank_data['Year'] = bank_data.index.astype(str).str[:4]

    savings_account = list(input.keys())[0]
    amount = bank_data.columns[1]

    bank_data = bank_data[bank_data['Categories'] != savings_account]
    bank_data = bank_data.groupby(['Year',
                                   'Month',
                                   'Categories']).agg({amount : 'sum'})

    for month in bank_data.index.get_level_values(1).astype(str).unique():
        for category in bank_data.index.get_level_values(2).unique():
            bank_data_final[month, category] = 0

    bank_data_final = pd.DataFrame(bank_data_final, index=[
                               'Placeholder']).T.sort_index()

    for year in bank_data.index.get_level_values(0).unique():
        bank_data_final[year] = bank_data.iloc[bank_data.index.get_level_values(
            0) == year].droplevel(0)

    for month in bank_data_final.index.get_level_values(0):
        total_list[month, '_Profit/Loss'] = bank_data_final.iloc[bank_data_final.index.get_level_values(0) == month].sum()

    totals = pd.DataFrame(total_list).T

    bank_data_final = pd.concat(
        [bank_data_final, totals]).sort_index().drop('Placeholder', axis=1)
    bank_data_final = bank_data_final.fillna(0)

    sheet_name = 'Total Overview.xlsx'
    folder = output
    full_path = folder + "\\" + sheet_name

    if sheet_name not in os.listdir(folder):
        wb = Workbook()
        wb.save(filename=full_path)

    book = load_workbook(full_path)
    writer = pd.ExcelWriter(full_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    try:
        book.remove_sheet(book['Sheet'])
    except KeyError:
        None

    bank_data_final.to_excel(writer, 'Overview')

    sheet = book['Overview']
    sheet.column_dimensions['A'].width = len(bank_data_final.index.get_level_values(0)[0]) * 1.5
    sheet.column_dimensions['B'].width = len(bank_data_final.index.get_level_values(1)[0]) * 2.0
    sheet.sheet_view.showGridLines = False

    writer.save()
